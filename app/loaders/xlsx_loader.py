"""
XlsxLoader — converts Excel workbooks to structured Markdown for RAG ingestion.

Problem with naive Excel → text:
  Flattening all sheets loses sheet boundaries and context.
  "Integration" rows mixed with "Support Matrix" rows → poor retrieval.
  Pipe tables embed poorly for natural language queries.

Our approach:
  1. Each sheet → its own H2 section (## SheetName)
     Chunker's header-aware splitting keeps rows within their sheet.

  2. Each row → prose sentence with column headers as context:
     "Module: Integration Hub | Type: REST API | Direction: Outbound"

  3. Sheet name embedded as H2 heading → becomes h2 breadcrumb in chunk
     metadata. ChunkContextualiser enriches:
     "This is from the Integration sheet of SPT_KnowledgeBase.xlsx..."

  4. Links preserved as Markdown inline links: [label](url)

  5. Empty rows and hidden/empty sheets skipped cleanly.

  6. Merged cells: loads without read_only to correctly resolve merge ranges.

  7. Header detection scans first 3 rows — handles title rows above headers.

  8. Columns beyond Z use AA, AB... Excel-style naming.

Known limitations:
  - .xls (old binary format) not supported — openpyxl is xlsx-only.
    SharePoint normalises most .xls to .xlsx on download.
  - Password-protected workbooks raise DocumentParsingError.
  - Very large sheets (>500 rows) are truncated — configurable via
    MAX_ROWS_PER_SHEET.

Output example:
  # SPT_KnowledgeBase

  ## Integration
  Module: Integration Hub | Interface Type: REST API | Direction: Outbound
  Module: Auth Service | Interface Type: SOAP | Direction: Inbound

  ## Support Matrix
  Component: Pricing Engine | Owner: Brad Jorgenson | SLA: 99.9%

  ## Knowledge Artifacts
  Artifact: SPT Runbook | Type: Documentation | Link: [View](https://...)
"""

import asyncio
import re
from io import BytesIO

import structlog

from app.loaders.base import BaseLoader, ParsedDocument
from app.connectors.base import RawDocument
from app.core.exceptions import DocumentParsingError

log = structlog.get_logger(__name__)

# Sheets with these names are metadata/TOC — skip regardless of content
_SKIP_SHEET_NAMES = {
    "sheet1",
    "sheet2",
    "sheet3",
    "toc",
    "table of contents",
    "contents",
    "index",
    "readme",
    "instructions",
    "cover",
    "coversheet",
    "changelog",
    "revision history",
    "change log",
}

# URL pattern for link detection in cell values
_URL_PATTERN = re.compile(r"https?://\S+", re.IGNORECASE)

# Max rows per sheet — avoid embedding massive data tables
_MAX_ROWS_PER_SHEET = 500

# How many leading rows to scan when detecting headers
_HEADER_SCAN_ROWS = 3


def _col_name(index: int) -> str:
    """
    Convert 0-based column index to Excel-style column name.
    0 → A, 25 → Z, 26 → AA, 27 → AB, 52 → BA ...
    Handles any number of columns correctly.
    """
    name = ""
    index += 1  # 1-based
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


class XlsxLoader(BaseLoader):
    """
    Loads Excel (.xlsx) workbooks into structured Markdown.
    Handles multi-sheet workbooks with mixed data structures.
    Uses openpyxl for parsing — no pandas dependency.
    """

    def __init__(self, llm=None) -> None:
        self._llm = llm
        log.info("loader.xlsx.initialised", llm_enrichment=llm is not None)

    def supported_mime_types(self) -> list[str]:
        return [
            # xlsx only — openpyxl cannot open .xls binary format
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ]

    def supported_structured_types(self) -> list[str]:
        return []

    async def load(self, raw_doc: RawDocument) -> ParsedDocument:
        """
        Parse Excel workbook into structured Markdown.
        Each sheet becomes a H2 section with rows as prose.
        """
        try:
            markdown, metadata = await asyncio.to_thread(
                self._parse_workbook,
                raw_doc.content,
                raw_doc.file_name,
            )
        except Exception as e:
            raise DocumentParsingError(
                f"Failed to parse Excel file '{raw_doc.file_name}': {e}"
            ) from e

        if not markdown.strip():
            log.warning(
                "loader.xlsx.empty",
                file=raw_doc.file_name,
                reason="no content extracted from any sheet",
            )
            return ParsedDocument(markdown_text="", extracted_metadata=metadata)

        log.info(
            "loader.xlsx.parsed",
            file=raw_doc.file_name,
            sheets=metadata.get("sheet_count", 0),
            rows=metadata.get("total_rows", 0),
        )
        return ParsedDocument(
            markdown_text=markdown,
            extracted_metadata=metadata,
        )

    # ── Core parsing ──────────────────────────────────────────────────────────

    def _parse_workbook(self, content: bytes, file_name: str) -> tuple[str, dict]:
        """Parse all visible sheets and return markdown + metadata."""
        import openpyxl

        # read_only=False required to correctly resolve merged cell ranges.
        # read_only=True is faster but merged_cells is unsupported and
        # silently returns empty values for merged cells — wrong for PwC docs.
        wb = openpyxl.load_workbook(
            BytesIO(content),
            read_only=False,
            data_only=True,  # get computed values, not formulas
        )

        sections = []
        total_rows = 0
        processed_sheets = []
        skipped_sheets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Skip hidden sheets — internal calculation/temp sheets
            if ws.sheet_state != "visible":
                skipped_sheets.append(sheet_name)
                log.debug(
                    "loader.xlsx.sheet_skipped",
                    sheet=sheet_name,
                    reason="hidden",
                )
                continue

            # Skip known metadata/TOC sheet names
            if sheet_name.strip().lower() in _SKIP_SHEET_NAMES:
                skipped_sheets.append(sheet_name)
                log.debug(
                    "loader.xlsx.sheet_skipped",
                    sheet=sheet_name,
                    reason="metadata_sheet",
                )
                continue

            section, row_count = self._parse_sheet(ws, sheet_name)

            if not section:
                skipped_sheets.append(sheet_name)
                log.debug(
                    "loader.xlsx.sheet_skipped",
                    sheet=sheet_name,
                    reason="empty",
                )
                continue

            sections.append(section)
            total_rows += row_count
            processed_sheets.append(sheet_name)
            log.debug(
                "loader.xlsx.sheet_parsed",
                sheet=sheet_name,
                rows=row_count,
            )

        wb.close()

        if not sections:
            return "", {
                "sheet_count": 0,
                "sheets": [],
                "skipped_sheets": skipped_sheets,
                "total_rows": 0,
                "file_name": file_name,
            }

        # H1 is the file stem (without extension)
        file_stem = file_name.rsplit(".", 1)[0] if "." in file_name else file_name
        parts = [f"# {file_stem}"]
        parts.extend(sections)
        markdown = "\n\n".join(parts)

        metadata = {
            "sheet_count": len(processed_sheets),
            "sheets": processed_sheets,
            "skipped_sheets": skipped_sheets,
            "total_rows": total_rows,
            "file_name": file_name,
        }

        return markdown, metadata

    def _parse_sheet(self, ws, sheet_name: str) -> tuple[str, int]:
        """
        Parse one sheet into a Markdown section.
        Returns (markdown_section, row_count).
        Returns ("", 0) if sheet is empty or yields no prose.
        """
        # Read all visible non-empty rows, resolve merged cells
        rows = self._read_rows(ws)

        if not rows:
            return "", 0

        # Scan first N rows to find the best header row
        # Handles: title rows above the real header row
        header_idx, headers = self._detect_headers(rows)

        # Data rows start after the header row (or from row 0 if no header)
        data_rows = rows[header_idx + 1 :] if header_idx >= 0 else rows

        if not data_rows:
            return "", 0

        # Convert each data row to a prose line
        prose_lines = []
        for row in data_rows[:_MAX_ROWS_PER_SHEET]:
            line = self._row_to_prose(row, headers)
            if line:
                prose_lines.append(line)

        if not prose_lines:
            return "", 0

        # ## SheetName heading → becomes h2 in chunk breadcrumb metadata
        section_lines = [f"## {sheet_name}", ""]
        section_lines.extend(prose_lines)

        return "\n".join(section_lines), len(prose_lines)

    def _read_rows(self, ws) -> list[list[str]]:
        """
        Read all rows, resolving merged cell values.
        Skips completely empty rows.
        """
        # Build merged cell map: (row, col) → value from merge's top-left cell
        # This requires read_only=False — merged_cells.ranges unavailable in read_only
        merged_values: dict[tuple[int, int], str] = {}
        try:
            for merge_range in ws.merged_cells.ranges:
                top_left_cell = ws.cell(merge_range.min_row, merge_range.min_col)
                value = self._cell_value(top_left_cell)
                for r in range(merge_range.min_row, merge_range.max_row + 1):
                    for c in range(merge_range.min_col, merge_range.max_col + 1):
                        merged_values[(r, c)] = value
        except (AttributeError, TypeError):
            pass

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            cells = []
            for col_idx, cell in enumerate(row, start=1):
                if (row_idx, col_idx) in merged_values:
                    cells.append(merged_values[(row_idx, col_idx)])
                else:
                    cells.append(self._cell_value(cell))

            # Skip completely empty rows
            if any(c.strip() for c in cells):
                rows.append(cells)

        return rows

    def _cell_value(self, cell) -> str:
        """
        Extract cell value as clean string.
        Preserves hyperlinks as Markdown [label](url).
        Converts raw URLs in cell text to [Link](url).
        """
        val = cell.value
        if val is None:
            return ""

        text = str(val).strip()
        if not text:
            return ""

        # Hyperlink attached to cell (Excel hyperlink dialog)
        try:
            if cell.hyperlink and cell.hyperlink.target:
                url = cell.hyperlink.target.strip()
                if url:
                    label = (
                        text
                        if text
                        and text.lower()
                        not in (url.lower(), "link", "click here", "here")
                        else "Link"
                    )
                    return f"[{label}]({url})"
        except AttributeError:
            pass

        # Raw URL value in cell
        if _URL_PATTERN.match(text):
            return f"[Link]({text})"

        return text

    def _detect_headers(
        self,
        rows: list[list[str]],
    ) -> tuple[int, list[str]]:
        """
        Scan first _HEADER_SCAN_ROWS rows to find the best header row.

        Strategy:
          For each candidate row, score it as a header:
            + Short strings (< 60 chars each) — column labels are brief
            + No pure numbers — headers are text not data
            + No URLs — headers don't contain links
            + Multiple non-empty cells — at least 2 columns
            - Long descriptive text — likely a title/subtitle row

          The highest-scoring row within the scan window is the header.
          If no row scores well → no header (use Column A, B, C...).

        Returns:
          (header_row_index, headers_list)
          header_row_index = -1 if no header detected
        """
        if not rows:
            return -1, []

        best_idx = -1
        best_score = 0

        scan_limit = min(_HEADER_SCAN_ROWS, len(rows))

        for i in range(scan_limit):
            row = rows[i]
            non_empty = [c for c in row if c.strip()]

            if len(non_empty) < 2:
                continue

            score = 0
            for c in non_empty:
                # Short cell → more likely a header label
                if len(c) <= 30:
                    score += 2
                elif len(c) <= 60:
                    score += 1
                else:
                    score -= 2  # long text → likely a title/subtitle

                # Pure number → definitely not a header
                if c.replace(".", "").replace(",", "").replace("-", "").isdigit():
                    score -= 3

                # URL → not a header
                if _URL_PATTERN.match(c):
                    score -= 3

                # Looks like a typical column label (Title Case or UPPER)
                if c.istitle() or c.isupper():
                    score += 1

            # Normalise by number of non-empty cells
            normalised = score / len(non_empty)

            if normalised > 0.5 and score > best_score:
                best_score = score
                best_idx = i

        if best_idx >= 0:
            header_row = rows[best_idx]
            headers = [
                c.strip() if c.strip() else _col_name(i)
                for i, c in enumerate(header_row)
            ]
            return best_idx, headers

        # No header detected — generate Column A, B, C...
        if rows:
            width = max(len(r) for r in rows[:scan_limit])
            headers = [_col_name(i) for i in range(width)]
        else:
            headers = []

        return -1, headers

    def _row_to_prose(self, row: list[str], headers: list[str]) -> str:
        """
        Convert a data row to a prose sentence.
        "Header1: Value1 | Header2: Value2 | Header3: Value3"

        Rules:
          - Skip empty cells
          - Skip cells with suspiciously long values (binary/formula noise)
          - If header is generic (none/null/-) just use the value
          - Handles rows wider than headers list
        """
        parts = []
        for i, value in enumerate(row):
            value = value.strip()
            if not value:
                continue
            if len(value) > 500:  # binary or formula noise
                continue

            # Get header — fallback to Column name if row wider than headers
            header = headers[i].strip() if i < len(headers) else _col_name(i)

            # Skip meaningless header names
            if not header or header.lower() in ("none", "null", "-", "n/a", ""):
                parts.append(value)
            else:
                parts.append(f"{header}: {value}")

        return " | ".join(parts) if parts else ""
